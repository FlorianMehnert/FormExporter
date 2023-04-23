import pathlib

import openpyxl
import pandas as pd
import xlrd
import yaml


class Pair:
    def __init__(self, one, two):
        self.one = one
        self.two = two


def fill_second_parameter(parameters: Pair, ls, index):
    bold_lines = get_header_line_numbers(ls)
    if parameters.one in ls[index]:
        ul = find_next_number(bold_lines, index)
        parameters.two = construct_content_of_lines(ls, ul, index)


def create_df_row(p: Pair):
    return p.two,


class Form:
    def __init__(self, h):
        self.vorname = Pair(h[0], None)  # titel, lokaler content
        self.nachname = Pair(h[1], None)
        self.bestelle_als = Pair(h[2], None)
        self.einrichtung = Pair(h[3], None)
        self.strasse = Pair(h[4], None)
        self.hausnummer = Pair(h[5], None)
        self.plz = Pair(h[6], None)
        self.stadt = Pair(h[7], None)
        self.email = Pair(h[8], None)
        self.telefon = Pair(h[9], None)
        self.erfahren = Pair(h[10], None)
        self.mitteilen = Pair(h[11], None)

    def get_dataframe(self):
        return [create_df_row(self.vorname),
                create_df_row(self.nachname),
                create_df_row(self.bestelle_als),
                create_df_row(self.einrichtung),
                create_df_row(self.strasse),
                create_df_row(self.hausnummer),
                create_df_row(self.plz),
                create_df_row(self.stadt),
                create_df_row(self.email),
                create_df_row(self.telefon),
                create_df_row(self.erfahren),
                create_df_row(self.mitteilen)]


def is_headline(l: str):
    l = l.strip()
    return l in headings


def get_header_line_numbers(lines: list):
    """
    this function finds all lines that contain a bold word marking a heading
    :return:
    """
    numbers = []
    for i, line in enumerate(lines):
        if is_headline(line):
            numbers.append(i)
    numbers.append(len(lines) + 1)
    return numbers


def find_next_number(number_list, number):
    for current_number in number_list:
        if number < current_number:
            return current_number


def construct_content_of_lines(lines: list[str], until, starting):
    content = ""
    for i, line in enumerate(lines):
        if until > i > starting:
            if i == until - 1:
                content += line.replace("\n", "")
            else:
                content += line
    return content


if __name__ == '__main__':
    file_name = ""
    table_name = ""
    sheet_name = ""
    headings = []

    with open("config.yml", "r") as stream:
        try:
            load = yaml.safe_load(stream)
            sheet_name = load['sheet']
            table_name = load['table']
            file_name = load['input']
            headings = load['headings']
        except yaml.YAMLError as exc:
            print(exc)

    if not file_name:
        print("Name der Eingabedatei fehlt")
        exit(1)
    if not table_name:
        print("Name der Tabelle fehlt")
        exit(1)
    if not sheet_name:
        print("Name des Tabellenblatts fehlt")
        exit(1)
    if not headings:
        print("Es wurden keine Tabellenspalten angegeben")
        exit(1)
    if len(headings) < 12:
        print("Es wurden zu wenig Tabellenspalten angeben")
        exit(1)

    form = Form(headings)
    if not pathlib.Path(file_name).exists():
        pathlib.Path(file_name).open("a+")
    with open(file_name, "r") as f:
        lines = f.readlines()
        for i, line in enumerate(lines):
            extra_parameters = (lines, i)
            fill_second_parameter(form.vorname, *extra_parameters)
            fill_second_parameter(form.nachname, *extra_parameters)
            fill_second_parameter(form.bestelle_als, *extra_parameters)
            fill_second_parameter(form.einrichtung, *extra_parameters)
            fill_second_parameter(form.telefon, *extra_parameters)
            fill_second_parameter(form.email, *extra_parameters)
            fill_second_parameter(form.stadt, *extra_parameters)
            fill_second_parameter(form.plz, *extra_parameters)
            fill_second_parameter(form.erfahren, *extra_parameters)
            fill_second_parameter(form.mitteilen, *extra_parameters)
            fill_second_parameter(form.strasse, *extra_parameters)
            fill_second_parameter(form.hausnummer, *extra_parameters)

    excel_file = pathlib.Path(table_name)
    if not excel_file.exists():
        wb = openpyxl.Workbook()
        wb.save(excel_file)

    with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df = pd.DataFrame.from_records(form.get_dataframe()).transpose()
        print(writer.sheets[sheet_name].max_row)
        df.to_excel(writer, sheet_name=sheet_name, startrow=writer.sheets[sheet_name].max_row, index=False,
                    header=False)
        writer._save()
